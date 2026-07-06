import csv
import hashlib
import sqlite3
import os
import re
import openpyxl

def generate_staff_id(name):
    """Creates a unique, consistent ID based on the staff name."""
    fingerprint = hashlib.md5(str(name).strip().lower().encode()).hexdigest()[:6].upper()
    return f"STF-{fingerprint}"

def extract_month_year(text):
    """
    Scans a string (like a file name or sheet name) for a Month and a Year.
    Returns the month number and the year if found, otherwise returns (None, None).
    """
    months = {
        'january': 1, 'jan': 1, 'february': 2, 'feb': 2, 'march': 3, 'mar': 3,
        'april': 4, 'apr': 4, 'may': 5, 'june': 6, 'jun': 6, 'july': 7, 'jul': 7,
        'august': 8, 'aug': 8, 'september': 9, 'sep': 9, 'october': 10, 'oct': 10,
        'november': 11, 'nov': 11, 'december': 12, 'dec': 12
    }
    
    text_lower = str(text).lower()
    found_month = None
    
    for month_name, month_num in months.items():
        if re.search(rf'\b{month_name}\b', text_lower):
            found_month = month_num
            break
            
    year_match = re.search(r'(201[0-9]|202[0-9]|203[0-5])', text)
    found_year = int(year_match.group(1)) if year_match else None
    
    if found_month and found_year:
        return found_month, found_year
    return None, None

def process_data_row(row, cursor, month, year, source_name):
    """A helper function to clean a single row of data and insert it."""
    def safe_str(val): return str(val).strip() if val is not None else ""
    
    if not row or len(row) < 15: 
        return 0
        
    staff_name = safe_str(row[0])
    ignore_list = ["STAFF NAME", "DG OFFICE", "LEG. SUPPORT SEVICES", "", "None"]
    
    if staff_name and staff_name.upper() not in ignore_list:
        try:
            designation = safe_str(row[3])
            net_salary_str = safe_str(row[-1]).replace(',', '')
            if not net_salary_str: return 0
            
            net_salary = float(net_salary_str)
            staff_id = generate_staff_id(staff_name)
            
            cursor.execute("""
                INSERT INTO historical_payroll 
                (staff_id, staff_name, designation, net_salary, payroll_month, payroll_year, source_file)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (staff_id, staff_name, designation, net_salary, month, year, source_name))
            
            return 1 
        except ValueError:
            return 0 
    return 0

def smart_batch_process(folder_path):
    conn = sqlite3.connect('hawea_payroll.db')
    cursor = conn.cursor()
    
    # THE BOUNCER: Tracks which (month, year) we have already processed
    processed_periods = set() 
    
    total_files_scanned = 0
    total_records_added = 0
    
    print(f"Oga, scanning the '{folder_path}' folder...\n")
    
    all_files = os.listdir(folder_path)
    
    # 1. PROCESS EXCEL FILES (.xlsx) FIRST (The VIPs)
    excel_files = [f for f in all_files if f.endswith('.xlsx')]
    for filename in excel_files:
        file_path = os.path.join(folder_path, filename)
        print(f"📂 Opening Excel Workbook: {filename}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                month, year = extract_month_year(sheet_name)
                
                if month and year:
                    period_key = (month, year)
                    if period_key in processed_periods:
                        print(f"   ⏩ Skipped sheet '{sheet_name}': We already have data for Month {month}, {year}.")
                        continue
                    
                    sheet = workbook[sheet_name]
                    sheet_records = 0
                    
                    for row in sheet.iter_rows(values_only=True):
                        source_tag = f"{filename} (Sheet: {sheet_name})"
                        sheet_records += process_data_row(row, cursor, month, year, source_tag)
                        
                    if sheet_records > 0:
                        processed_periods.add(period_key)
                        total_records_added += sheet_records
                        print(f"   ✅ Processed sheet '{sheet_name}' -> Saved {sheet_records} records.")
        except Exception as e:
             print(f"   ❌ Error reading Excel file {filename}: {e}")
             
        total_files_scanned += 1

    # 2. PROCESS CSV FILES SECOND
    csv_files = [f for f in all_files if f.endswith('.csv')]
    for filename in csv_files:
        file_path = os.path.join(folder_path, filename)
        month, year = extract_month_year(filename)
        
        if month and year:
            period_key = (month, year)
            if period_key in processed_periods:
                print(f"⏩ Skipped CSV '{filename}': We already have data for Month {month}, {year} from an Excel sheet.")
                continue
            
            file_records = 0
            try:
                with open(file_path, mode='r', encoding='utf-8') as file:
                    reader = csv.reader(file)
                    for row in reader:
                        file_records += process_data_row(row, cursor, month, year, filename)
                        
                if file_records > 0:
                    processed_periods.add(period_key)
                    total_records_added += file_records
                    print(f"✅ Processed CSV '{filename}' -> Saved {file_records} records.")
            except Exception as e:
                print(f"❌ Error reading CSV {filename}: {e}")
        else:
            print(f"⚠️ Skipped CSV '{filename}': Could not detect a clear Month and Year in the name.")
            
        total_files_scanned += 1

    conn.commit()
    conn.close()
    
    print("\n" + "="*50)
    print("🎯 SMART BATCH PROCESSING COMPLETE!")
    print(f"Total Files Scanned: {total_files_scanned}")
    print(f"Unique Months Reconciled: {len(processed_periods)}")
    print(f"Total Staff Records Saved: {total_records_added}")
    print("="*50)

# --- Run Execution ---
if __name__ == "__main__":
    target_folder = "legacy_payroll_files"
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)
        print(f"Created '{target_folder}'. Drop your files inside and run again!")
    else:
        smart_batch_process(target_folder)