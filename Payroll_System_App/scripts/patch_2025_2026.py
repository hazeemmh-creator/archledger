import os
import openpyxl
import pandas as pd

# --- CONFIGURATION ---
SOURCE_FOLDERS = ["./Extracted_By_Year/2025", "./Extracted_By_Year/2026"]
OUTPUT_FILE = "Fixed_2025_2026_Payroll.csv"

# THE HARDCODED GPS MAP
COLUMN_MAP = {
    "B": "STAFF NAME",
    "C": "DESIGNATION",
    "D": "GRADE LEVEL",
    "E": "BASIC SALARY",
    "F": "ALLCE/ AREARS",
    "G": "RESEARCH ALLOWANCE",
    "H": "HARDSHIP ALLOWANCE",
    "I": "LEGISLATIVE DUTY ALLOWANCE",
    "J": "GROSS SALARY",
    "K": "AUCTION",
    "L": "COOP 1 CONTR/SPEC SAVINGS",
    "M": "COOP 1 LOAN RECOVERY",
    "N": "COOP 2 CONTR/SPEC SAVINGS",
    "O": "COOP 2 LOAN RECOVERY",
    "P": "NHF DED.",
    "Q": "PAYE",
    "R": "EMPLOYEE PENSIONS",
    "U": "TOTAL DEDUCTION",  # <-- Your exact coordinate!
    "V": "PREVIOUS MONTH",
    "W": "NET PAY"           # <-- Your exact coordinate!
}

MONTHS = {
    "january": "1", "february": "2", "march": "3",
    "april": "4", "may": "5", "june": "6", "july": "7",
    "august": "8", "september": "9", "october": "10",
    "november": "11", "december": "12"
}

def gps_extract_batch():
    print("Oga, initiating GPS Override for 2025 and 2026...")
    all_data = []

    for folder in SOURCE_FOLDERS:
        if not os.path.exists(folder):
            print(f"Skipping {folder} - Not found.")
            continue
            
        year = os.path.basename(folder)
        
        for file in os.listdir(folder):
            if not file.endswith(('.xlsx', '.xls')): continue
            filepath = os.path.join(folder, file)
            
            print(f"🎯 Dropping GPS pins on: {file}")
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
                
                extracted_data = []
                for row_num in range(8, ws.max_row + 1):
                    staff_name = ws[f"B{row_num}"].value
                    
                    if not staff_name or str(staff_name).upper().strip() in ['NONE', 'NAME']:
                        continue
                    if "TOTAL" in str(staff_name).upper():
                        continue

                    row_data = {}
                    for col_letter, db_column_name in COLUMN_MAP.items():
                        cell_value = ws[f"{col_letter}{row_num}"].value
                        row_data[db_column_name] = cell_value
                        
                    target_month = "Unknown"
                    for m_key, m_val in MONTHS.items():
                        if m_key in file.lower():
                            target_month = m_val
                            break
                            
                    row_data["Payroll_Month"] = target_month
                    row_data["Payroll_Year"] = year
                        
                    extracted_data.append(row_data)

                if extracted_data:
                    df = pd.DataFrame(extracted_data)
                    all_data.append(df)
                    
            except Exception as e:
                print(f"❌ Error on {file}: {e}")

    if all_data:
        print("\nForging the Fixed Payload...")
        master_df = pd.concat(all_data, ignore_index=True)
        master_df.to_csv(OUTPUT_FILE, index=False)
        print(f"Gbosa! Clean file saved as '{OUTPUT_FILE}'.")
    else:
        print("⚠️ Failed to extract data.")

if __name__ == "__main__":
    gps_extract_batch()